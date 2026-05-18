"""
╔══════════════════════════════════════════════════════════════════════╗
║           GOOFY SCREENER — PHASE 8                                   ║
║           Paper Trading Framework                                    ║
║                                                                      ║
║   Inherits Phase 7 (ML signal layer) — adds a persistent paper      ║
║   trading log that records every TRADE signal at today's price,      ║
║   marks open positions to market, and closes them after 20 days.    ║
║                                                                      ║
║   Three new Excel tabs:                                              ║
║       📋 Open Positions  — live paper trades + unrealised P&L       ║
║       📈 Trade History   — closed trades with actual returns         ║
║       📊 P8 Performance  — win rate, Sharpe, breakdown by tier      ║
║                                                                      ║
║   Run:  python3 goofy_screener_phase8.py                            ║
║   Args: --market US | ASX | JPX | ALL (default: ALL)               ║
╚══════════════════════════════════════════════════════════════════════╝
"""

import yfinance as yf
import numpy as np
import pandas as pd
import datetime as dt
import os, sys, warnings, argparse
warnings.filterwarnings("ignore")

# ── Phase 5–7 layers ─────────────────────────────────────────────────────────
from goofy_screener_phase5 import (
    UNIVERSE_MAP, STRATEGY_FNS, STRATEGY_GRIDS,
    TRAIN_START, TRAIN_END, TEST_END, MIN_ROWS,
    TARGET_VOL, KELLY_FRACTION, OUTPUT_DIR,
    compute_metrics, score_asset, compute_today_verdict,
    apply_sheet_formatting, _style_header,
    LONG_TO_SHORT,
)
from position_sizer import compute_trade_stats, recommend_size
from regime_detector import load_asset_gates
from portfolio_builder import (
    compute_correlation_matrix, find_clusters,
    adjust_for_correlation, portfolio_metrics,
    cluster_label, CORR_THRESHOLD,
)
from ml_signal import (
    engineer_features, build_ml_model, get_current_score,
    ml_gate, combined_verdict, XGB_AVAILABLE,
    ML_PASS_THRESH, LOOKFORWARD,
)
from goofy_screener_phase7 import screen_market_p7, run_ml_layer, write_excel_phase7

# ── Phase 8 paper trading layer ───────────────────────────────────────────────
from paper_trader import (
    log_new_signals, mark_to_market, get_performance,
    get_open_positions, get_closed_trades, HOLD_DAYS,
)

try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    EXCEL_FORMAT = True
except ImportError:
    EXCEL_FORMAT = False

OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "screener_output")


# ══════════════════════════════════════════════════════════════════════════════
#  PHASE 8 EXCEL WRITER
# ══════════════════════════════════════════════════════════════════════════════

def _write_open_positions_sheet(wb, open_df: pd.DataFrame, today: str):
    ws = wb.create_sheet(title="📋 Open Positions")

    ws.append(["PHASE 8 — OPEN PAPER POSITIONS"])
    ws.append([f"Run: {today}  |  Each position closes after {HOLD_DAYS} trading days"])
    ws.append([])

    if open_df.empty:
        ws.append(["No open positions."])
        return

    cols = ["asset", "market", "entry_date", "entry_price", "current_price",
            "unrealised_pnl_pct", "days_held", "hold_days_target",
            "size_pct", "ml_score", "val_auc", "tier", "strategy"]
    cols = [c for c in cols if c in open_df.columns]
    header_map = {
        "asset": "Asset", "market": "Market", "entry_date": "Entry Date",
        "entry_price": "Entry Price", "current_price": "Current Price",
        "unrealised_pnl_pct": "Unreal. P&L %", "days_held": "Days Held",
        "hold_days_target": "Hold Target", "size_pct": "Size %",
        "ml_score": "ML Score", "val_auc": "Val AUC",
        "tier": "Tier", "strategy": "Strategy",
    }
    ws.append([header_map.get(c, c) for c in cols])
    _style_header(ws, row_num=4)

    for _, row in open_df[cols].iterrows():
        ws.append([row[c] for c in cols])

    if EXCEL_FORMAT:
        pnl_col_idx = cols.index("unrealised_pnl_pct") + 1 if "unrealised_pnl_pct" in cols else None
        for r in ws.iter_rows(min_row=5, max_row=ws.max_row):
            for cell in r:
                if pnl_col_idx and cell.column == pnl_col_idx and cell.value is not None:
                    try:
                        v = float(cell.value)
                        cell.fill = PatternFill("solid", fgColor="27AE60" if v >= 0 else "E74C3C")
                        cell.font = Font(bold=True, color="FFFFFF", size=10)
                    except Exception:
                        pass

    ws.auto_filter.ref = f"A4:{get_column_letter(len(cols))}{ws.max_row}"
    ws.freeze_panes = "A5"
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 28)


def _write_trade_history_sheet(wb, closed_df: pd.DataFrame, today: str):
    ws = wb.create_sheet(title="📈 Trade History")

    ws.append(["PHASE 8 — CLOSED TRADE HISTORY"])
    ws.append([f"Run: {today}  |  20-day hold, no transaction costs"])
    ws.append([])

    if closed_df.empty:
        ws.append([f"No closed trades yet — check back in {HOLD_DAYS} trading days (≈ 1 calendar month)."])
        return

    cols = ["asset", "market", "entry_date", "entry_price", "exit_date", "exit_price",
            "pnl_pct", "win", "days_held", "size_pct", "ml_score", "val_auc", "tier", "strategy"]
    cols = [c for c in cols if c in closed_df.columns]
    header_map = {
        "asset": "Asset", "market": "Market", "entry_date": "Entry Date",
        "entry_price": "Entry Price", "exit_date": "Exit Date",
        "exit_price": "Exit Price", "pnl_pct": "P&L %", "win": "Win?",
        "days_held": "Days Held", "size_pct": "Size %",
        "ml_score": "ML Score", "val_auc": "Val AUC",
        "tier": "Tier", "strategy": "Strategy",
    }
    ws.append([header_map.get(c, c) for c in cols])
    _style_header(ws, row_num=4)

    sorted_df = closed_df.sort_values("exit_date", ascending=False) if "exit_date" in closed_df.columns else closed_df
    for _, row in sorted_df[cols].iterrows():
        ws.append([row[c] for c in cols])

    if EXCEL_FORMAT:
        pnl_idx = cols.index("pnl_pct") + 1 if "pnl_pct" in cols else None
        win_idx = cols.index("win") + 1 if "win" in cols else None
        for r in ws.iter_rows(min_row=5, max_row=ws.max_row):
            for cell in r:
                if pnl_idx and cell.column == pnl_idx and cell.value is not None:
                    try:
                        v = float(cell.value)
                        cell.fill = PatternFill("solid", fgColor="27AE60" if v >= 0 else "E74C3C")
                        cell.font = Font(bold=True, color="FFFFFF", size=10)
                    except Exception:
                        pass
                if win_idx and cell.column == win_idx:
                    if cell.value is True:
                        cell.value = "✓"; cell.font = Font(color="27AE60", bold=True)
                    elif cell.value is False:
                        cell.value = "✗"; cell.font = Font(color="E74C3C", bold=True)

    ws.auto_filter.ref = f"A4:{get_column_letter(len(cols))}{ws.max_row}"
    ws.freeze_panes = "A5"
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 28)


def _write_performance_sheet(wb, perf: dict, today: str):
    ws = wb.create_sheet(title="📊 P8 Performance")

    ws.append(["PHASE 8 — PAPER TRADING PERFORMANCE"])
    ws.append([f"Run: {today}"])
    ws.append([])

    def row(label, value):
        ws.append([label, value])

    row("Closed trades",      perf.get("n_closed", 0))
    row("Open positions",     perf.get("n_open", 0))

    if perf.get("n_closed", 0) == 0:
        ws.append([])
        ws.append([perf.get("message", "No closed trades yet.")])
        ws["A1"].font = Font(bold=True, size=14)
        return

    ws.append([])
    row("Win rate",           f"{perf['win_rate']:.1%}")
    row("Avg P&L per trade",  f"{perf['avg_pnl_pct']:.2f}%")
    row("Median P&L",         f"{perf['median_pnl_pct']:.2f}%")
    row("Std dev P&L",        f"{perf['std_pnl_pct']:.2f}%")
    row("Annualised Sharpe",  perf.get("annualised_sharpe", "—"))
    ws.append([])
    row("Best trade",  f"{perf['best_trade']['asset']}  {perf['best_trade']['pnl_pct']:+.2f}%  ({perf['best_trade']['entry_date']})")
    row("Worst trade", f"{perf['worst_trade']['asset']}  {perf['worst_trade']['pnl_pct']:+.2f}%  ({perf['worst_trade']['entry_date']})")

    if perf.get("by_tier"):
        ws.append([])
        ws.append(["By Tier", "Count", "Win Rate", "Avg P&L %"])
        for tier, stats in sorted(perf["by_tier"].items()):
            ws.append([tier, stats["n"], f"{stats['win_rate']:.1%}", f"{stats['avg_pnl_pct']:.2f}%"])

    if perf.get("by_market"):
        ws.append([])
        ws.append(["By Market", "Count", "Win Rate", "Avg P&L %"])
        for mkt, stats in sorted(perf["by_market"].items()):
            ws.append([mkt, stats["n"], f"{stats['win_rate']:.1%}", f"{stats['avg_pnl_pct']:.2f}%"])

    if EXCEL_FORMAT:
        ws["A1"].font = Font(bold=True, size=14, color="1C2833")
        ws["A2"].font = Font(italic=True, size=10, color="555555")
        ws.column_dimensions["A"].width = 24
        ws.column_dimensions["B"].width = 28


def write_excel_phase8(all_results: dict, today: str, p6: dict,
                       open_df: pd.DataFrame, closed_df: pd.DataFrame,
                       perf: dict) -> str:
    """Build the Phase 8 Excel: Phase 7 tabs + three new P8 tabs."""
    from openpyxl import load_workbook

    # Write Phase 7 Excel first
    p7_path = write_excel_phase7(all_results, today, p6)

    # Reopen and append P8 tabs
    wb = load_workbook(p7_path)

    _write_open_positions_sheet(wb, open_df, today)
    _write_trade_history_sheet(wb, closed_df, today)
    _write_performance_sheet(wb, perf, today)

    # Rename and save as Phase 8
    fname = f"Goofy_Phase8_{today}.xlsx"
    path  = os.path.join(OUTPUT_DIR, fname)
    wb.save(path)

    # Remove Phase 7 file (Phase 8 is the full report)
    try:
        os.remove(p7_path)
    except Exception:
        pass

    return path


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="Goofy Phase 8 — Paper trading framework")
    parser.add_argument("--market", choices=["US", "ASX", "JPX", "ALL"], default="ALL")
    args  = parser.parse_args()
    today = dt.datetime.now().strftime("%Y-%m-%d")

    markets_to_run = ["US", "ASX", "JPX"] if args.market == "ALL" else [args.market]

    print(f"""
╔══════════════════════════════════════════════════════════════════════╗
║  GOOFY SCREENER — PHASE 8  |  {today}                    ║
║  Strategy + Regime + Kelly + Correlation + ML + Paper Trading       ║
║  Markets: {', '.join(markets_to_run):52} ║
║  Paper trades: {HOLD_DAYS}-day hold, logged to paper_trades/trades_log.json  ║
╚══════════════════════════════════════════════════════════════════════╝""")

    load_asset_gates()

    # ── Step 1: Download ──────────────────────────────────────────────────────
    all_assets = []
    for m in markets_to_run:
        all_assets.extend(UNIVERSE_MAP[m])
    seen = set(); unique_assets = []
    for a in all_assets:
        if a not in seen:
            unique_assets.append(a); seen.add(a)

    print(f"\n[1/6] Downloading {len(unique_assets)} assets...\n")
    price_data = {}; ohlc_data = {}
    for asset in unique_assets:
        try:
            raw = yf.download(asset, start=TRAIN_START, end=TEST_END,
                              auto_adjust=True, progress=False)
            if not raw.empty and len(raw) >= MIN_ROWS:
                if isinstance(raw.columns, pd.MultiIndex):
                    raw.columns = raw.columns.get_level_values(0)
                close = raw["Close"].squeeze()
                if isinstance(close, pd.DataFrame):
                    close = close.iloc[:, 0]
                price_data[asset] = close
                if {"High", "Low", "Close"}.issubset(set(raw.columns)):
                    ohlc_data[asset] = raw[["High", "Low", "Close"]].copy()
                print(f"  ✓ {asset:14} | {len(raw)} rows")
            else:
                print(f"  ✗ {asset:14} | skipped")
        except Exception as e:
            print(f"  ✗ {asset:14} | error: {e}")

    # ── Step 2: Phase 5 screening ─────────────────────────────────────────────
    print(f"\n[2/6] Phase 5 screening (strategy + regime + sizing)...\n")
    all_results = {}
    for m in markets_to_run:
        all_results[m] = screen_market_p7(m, UNIVERSE_MAP[m], price_data, ohlc_data)

    all_dfs = [df for df in all_results.values() if not df.empty]
    if not all_dfs:
        print("  No results."); return
    combined = pd.concat(all_dfs, ignore_index=True)

    # ── Step 3: Phase 6 portfolio construction ───────────────────────────────
    print(f"\n[3/6] Phase 6 — correlation matrix + cluster sizing...\n")
    from goofy_screener_phase6 import run_portfolio_construction, attach_phase6_columns
    p6 = run_portfolio_construction(combined, price_data)
    all_results = attach_phase6_columns(all_results, p6)
    combined = pd.concat([df for df in all_results.values() if not df.empty],
                         ignore_index=True)

    # ── Step 4: Phase 7 ML layer ──────────────────────────────────────────────
    print(f"\n[4/6] Phase 7 — XGBoost ML signal layer...\n")
    all_results = run_ml_layer(all_results, price_data, ohlc_data)
    combined = pd.concat([df for df in all_results.values() if not df.empty],
                         ignore_index=True)

    # Attach Adj Size % to combined (needed for paper trader)
    adj_sizes = p6.get("adj_sizes", {})
    combined["Adj Size %"] = combined["Asset"].map(adj_sizes)

    # ── Step 5: Phase 8 paper trading ────────────────────────────────────────
    print(f"\n[5/6] Phase 8 — paper trading update...\n")

    # Mark existing positions to market first
    still_open, newly_closed = mark_to_market(price_data, run_date=today)
    if newly_closed:
        print(f"  ✅ Closed {len(newly_closed)} position(s) after {HOLD_DAYS} days:")
        for t in newly_closed:
            sign = "+" if (t.get("pnl_pct") or 0) >= 0 else ""
            tag  = " [STOP]" if t.get("exit_reason") == "STOP_LOSS" else ""
            print(f"     {t['asset']:14} → {sign}{t.get('pnl_pct', '?'):.2f}%{tag}")
    else:
        print(f"  — No positions expired today.")

    # Log new TRADE signals
    new_entries = log_new_signals(combined, price_data, run_date=today)
    if new_entries:
        print(f"\n  📋 Logged {len(new_entries)} new paper trade(s):")
        for t in new_entries:
            print(f"     {t['asset']:14} @ {t['entry_price']:.4f}  "
                  f"[{t['tier']}]  ML={t.get('ml_score', '—')}%  "
                  f"Size={t['size_pct']:.1f}%")
    else:
        print(f"\n  — No new TRADE signals to log.")

    # Load current state
    open_df   = get_open_positions()
    closed_df = get_closed_trades()
    perf      = get_performance()

    # Print summary
    print(f"\n{'═'*72}")
    print(f"  PHASE 8 PAPER TRADING STATUS  |  {today}")
    print(f"{'═'*72}")
    print(f"  📋 Open positions : {perf.get('n_open', 0)}")
    print(f"  📈 Closed trades  : {perf.get('n_closed', 0)}"
          + (f"  ({perf['n_stopped']} stopped early)" if perf.get('n_stopped') else ""))
    if perf.get("n_closed", 0) > 0:
        print(f"  Win rate          : {perf['win_rate']:.1%}")
        print(f"  Avg P&L per trade : {perf['avg_pnl_pct']:+.2f}%")
        if perf.get("annualised_sharpe"):
            print(f"  Annualised Sharpe : {perf['annualised_sharpe']:.2f}")
    else:
        print(f"  → {perf.get('message', '')}")

    # Print open positions table
    if not open_df.empty:
        print(f"\n  ── Open positions ──")
        print(f"  {'Asset':14} {'Entry':12} {'Price':10} {'Curr':10} {'P&L%':8} {'Days':5} {'Tier'}")
        print(f"  {'─'*70}")
        for _, r in open_df.iterrows():
            pnl = r.get("unrealised_pnl_pct")
            pnl_str = f"{pnl:+.2f}%" if pnl is not None else "  —"
            curr = r.get("current_price")
            curr_str = f"{curr:.2f}" if curr else "—"
            days = r.get("days_held", 0)
            print(f"  {r['asset']:14} {r['entry_date']:12} {r['entry_price']:>10.4f} "
                  f"{curr_str:>10} {pnl_str:>8} {days:>5}  {r.get('tier','')}")

    # ── Step 6: Save Excel ────────────────────────────────────────────────────
    print(f"\n[6/6] Saving Phase 8 report...\n")
    xlsx_path = write_excel_phase8(all_results, today, p6, open_df, closed_df, perf)
    print(f"  Report saved → {os.path.basename(xlsx_path)}")
    print(f"  Full path:   {xlsx_path}")
    print(f"\n{'═'*72}\n")


if __name__ == "__main__":
    main()
